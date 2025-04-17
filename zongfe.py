import os
import glob
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 定义总表文件名
SUMMARY_FILE = '活动总分汇总表.xlsx'

def process_activity_data(file_pattern='S*.xlsx', update_mode=True):
    """
    处理所有活动数据文件并生成或更新总表，仅使用openpyxl库
    
    参数:
    file_pattern (str): 文件匹配模式，默认为'S*.xlsx'
    update_mode (bool): 是否使用更新模式（如果为True，将基于现有总表添加新数据）
    """
    # 获取所有匹配的文件并按名称排序
    files = sorted(glob.glob(file_pattern))
    
    if not files:
        print(f"未找到匹配 '{file_pattern}' 的文件")
        return
    
    print(f"找到以下文件: {files}")
    
    # 检查是否存在现有的总表，以及是否需要更新
    existing_data = {}
    existing_periods = []
    if update_mode and os.path.exists(SUMMARY_FILE):
        print(f"检测到现有总表文件: {SUMMARY_FILE}，将在此基础上更新")
        existing_data, existing_periods = read_existing_summary()
    
    # 用于存储所有数据的字典
    all_data = existing_data.copy()  # 复制现有数据（如果有）
    
    # 用于存储所有参与者ID的集合
    all_participants = set()
    if existing_data:
        all_participants = set(existing_data.keys())
    
    # 用于存储参与者信息的字典 (学号 -> [姓名, 手机])
    participant_info = {}
    for student_id, data in existing_data.items():
        participant_info[student_id] = data['info']
    
    # 处理需要添加的新文件
    processed_files = []
    for file in files:
        # 从文件名中提取期数
        period = file.replace('S', '').replace('.xlsx', '')
        period_key = f'S{period}'
        
        # 如果当前期数已经在总表中，则跳过（除非强制更新）
        if period_key in existing_periods and update_mode:
            print(f"期次 {period_key} 已存在于总表中，跳过文件 {file}")
            continue
        
        try:
            # 读取Excel文件
            wb = load_workbook(filename=file, read_only=True)
            ws = wb.active
            
            # 获取表头
            headers = []
            for cell in next(ws.rows):
                headers.append(cell.value)
            
            # 检查必要的列是否存在
            required_columns = ['年级专业班级姓名', '手机号码', '学号', '总分']
            missing_columns = [col for col in required_columns if col not in headers]
            
            if missing_columns:
                print(f"警告: 文件 {file} 缺少以下必要列: {missing_columns}")
                continue
            
            # 获取必要列的索引
            name_idx = headers.index('年级专业班级姓名')
            phone_idx = headers.index('手机号码')
            id_idx = headers.index('学号')
            score_idx = headers.index('总分')
            
            # 跳过表头，读取数据行
            row_count = 0
            for row in list(ws.rows)[1:]:
                row_count += 1
                # 确保行有足够的单元格
                if len(row) <= max(name_idx, phone_idx, id_idx, score_idx):
                    print(f"警告: 第 {row_count+1} 行数据不完整，已跳过")
                    continue
                
                # 读取单元格数据
                name = row[name_idx].value
                phone = row[phone_idx].value
                student_id = str(row[id_idx].value)  # 转为字符串确保一致性
                score = row[score_idx].value if row[score_idx].value is not None else 0
                
                # 跳过没有学号的记录
                if not student_id:
                    print(f"警告: 文件 {file} 中第 {row_count+1} 行缺少学号，已跳过")
                    continue
                
                # 更新参与者ID集合
                all_participants.add(student_id)
                
                # 存储学生信息
                if student_id not in participant_info:
                    participant_info[student_id] = [name, phone]
                
                # 如果学生不在all_data中，添加一个空字典
                if student_id not in all_data:
                    all_data[student_id] = {'info': [name, phone], 'scores': {}}
                
                # 存储分数
                all_data[student_id]['scores'][period_key] = score
            
            processed_files.append(file)
            print(f"成功处理文件 {file}，当前总表包含 {len(all_participants)} 名参与者")
            wb.close()
        
        except Exception as e:
            print(f"读取文件 {file} 时出错: {e}")
    
    if not processed_files and not existing_data:
        print("没有成功读取任何文件数据")
        return
    elif not processed_files and existing_data:
        print("没有新的文件需要处理，总表保持不变")
        return
    
    # 获取所有期次（包括现有的和新添加的）
    all_periods = sorted(set(existing_periods + [
        f'S{file.replace("S", "").replace(".xlsx", "")}' 
        for file in processed_files
    ]))
    
    # 创建新的工作簿用于输出
    result_wb = Workbook()
    result_ws = result_wb.active
    result_ws.title = "总分汇总"
    
    # 添加表头
    headers = ['年级专业班级姓名', '手机号码', '学号']
    headers.extend(all_periods)
    headers.append('总分')
    
    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        result_ws.cell(row=1, column=col_idx, value=header)
    
    # 为每个参与者添加数据行
    result_data = []  # 用于排序的临时数据存储
    
    for student_id in all_participants:
        student_data = all_data.get(student_id, {})
        student_info = participant_info.get(student_id, ['未知', '未知'])
        student_scores = student_data.get('scores', {}) if isinstance(student_data, dict) else {}
        
        # 计算总分并创建行数据
        row_data = [student_info[0], student_info[1], student_id]
        total_score = 0
        
        for period in all_periods:
            score = student_scores.get(period, 0)
            total_score += score
            row_data.append(score)
        
        row_data.append(total_score)
        result_data.append(row_data)
    
    # 按总分降序排序
    result_data.sort(key=lambda x: x[-1], reverse=True)
    
    # 写入排序后的数据
    for row_idx, row_data in enumerate(result_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            result_ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 美化输出文件
    beautify_excel(result_ws, all_periods)
    
    # 保存结果
    result_wb.save(SUMMARY_FILE)
    
    print(f"总表已生成/更新: {SUMMARY_FILE}")
    print(f"- 包含 {len(all_participants)} 名参与者")
    print(f"- 包含 {len(all_periods)} 期活动数据")
    
    return SUMMARY_FILE


def read_existing_summary():
    """
    读取现有的总表数据
    
    返回:
    tuple: (数据字典, 期次列表)
    """
    try:
        wb = load_workbook(filename=SUMMARY_FILE, read_only=True)
        ws = wb.active
        
        # 获取表头
        headers = []
        for cell in next(ws.rows):
            headers.append(cell.value)
        
        # 查找必要列的索引
        try:
            name_idx = headers.index('年级专业班级姓名')
            phone_idx = headers.index('手机号码')
            id_idx = headers.index('学号')
            total_idx = headers.index('总分')
        except ValueError as e:
            print(f"现有总表缺少必要的列: {e}")
            return {}, []
        
        # 提取期次列
        period_columns = []
        for i, header in enumerate(headers):
            if header.startswith('S') and i not in [name_idx, phone_idx, id_idx, total_idx]:
                period_columns.append(header)
        
        # 读取所有学生数据
        student_data = {}
        
        for row in list(ws.rows)[1:]:  # 跳过表头
            # 确保行有足够的单元格
            if len(row) <= max(name_idx, phone_idx, id_idx, total_idx):
                continue
            
            name = row[name_idx].value
            phone = row[phone_idx].value
            student_id = str(row[id_idx].value)
            
            if not student_id:
                continue
            
            # 创建学生记录
            student_data[student_id] = {
                'info': [name, phone],
                'scores': {}
            }
            
            # 读取各期分数
            for period in period_columns:
                period_idx = headers.index(period)
                if period_idx < len(row):
                    score = row[period_idx].value
                    student_data[student_id]['scores'][period] = score if score is not None else 0
        
        wb.close()
        return student_data, period_columns
    
    except Exception as e:
        print(f"读取现有总表时出错: {e}")
        return {}, []


def beautify_excel(ws, period_columns):
    """
    美化Excel工作表的格式
    
    参数:
    ws (Worksheet): 需要美化的工作表
    period_columns (list): 期数列的列表
    """
    # 定义样式
    header_fill = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid")
    period_fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
    total_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    zero_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    
    header_font = Font(bold=True)
    center_aligned = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )
    
    # 获取表头
    headers = [cell.value for cell in list(ws.rows)[0]]
    
    # 应用表头样式
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_aligned
        cell.border = border
    
    # 设置列宽
    ws.column_dimensions[get_column_letter(1)].width = 30  # 年级专业班级姓名
    ws.column_dimensions[get_column_letter(2)].width = 15  # 手机号码
    ws.column_dimensions[get_column_letter(3)].width = 15  # 学号
    
    # 为各期分数和总分列设置样式
    period_indices = []
    
    # 查找各期分数和总分列的索引
    for col_idx, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_idx)
        
        # 设置每列的基本样式
        ws.column_dimensions[column_letter].width = 12
        
        # 为期数列应用特殊样式
        if header in period_columns:
            for row_idx in range(2, ws.max_row + 1):
                period_cell = ws.cell(row=row_idx, column=col_idx)
                period_cell.fill = period_fill
                period_cell.alignment = center_aligned
                period_cell.border = border
                
                # 标记0分（未参加）的单元格
                if period_cell.value == 0:
                    period_cell.fill = zero_fill
            
            period_indices.append(col_idx)
        
        # 为总分列应用特殊样式
        elif header == "总分":
            for row_idx in range(2, ws.max_row + 1):
                total_cell = ws.cell(row=row_idx, column=col_idx)
                total_cell.fill = total_fill
                total_cell.alignment = center_aligned
                total_cell.font = header_font
                total_cell.border = border
    
    # 为所有其他单元格添加边框
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if col not in period_indices and headers[col-1] != "总分":
                cell = ws.cell(row=row, column=col)
                cell.border = border
    
    # 冻结首行
    ws.freeze_panes = "A2"


def print_usage():
    """打印使用说明"""
    print("\n活动成绩汇总程序 - 使用说明")
    print("=" * 50)
    print("本程序可以读取多个活动期次的Excel文件，生成或更新总分汇总表。")
    print("\n基本用法:")
    print("  1. 将所有期次的Excel文件(S1.xlsx, S2.xlsx等)放在与程序相同目录下")
    print("  2. 运行程序，自动生成或更新'活动总分汇总表.xlsx'")
    print("\n文件要求:")
    print("  - 每个输入文件必须包含以下列: 年级专业班级姓名, 手机号码, 学号, 总分")
    print("  - 文件命名应遵循'S数字.xlsx'格式，例如S1.xlsx, S2.xlsx等")
    print("\n更新模式:")
    print("  - 默认情况下，程序会检测现有的总表文件，并在其基础上添加新期次的数据")
    print("  - 如需从头重新生成总表，请选择选项2")
    print("=" * 50)


if __name__ == "__main__":
    try:
        print_usage()
        print("\n请选择操作模式:")
        print("  1. 更新模式 - 如果存在总表，则在其基础上更新（推荐）")
        print("  2. 重新生成 - 忽略现有总表，重新生成完整的总表")
        
        choice = input("请输入选项(1或2): ").strip()
        update_mode = True if choice != "2" else False
        
        if update_mode:
            print("\n已选择更新模式 - 将在现有总表的基础上添加新数据")
        else:
            print("\n已选择重新生成模式 - 将创建全新的总表")
        
        output_file = process_activity_data(update_mode=update_mode)
        if output_file:
            print(f"\n处理完成。您可以打开 {output_file} 查看结果。")
    except Exception as e:
        print(f"\n程序执行过程中发生错误: {e}")